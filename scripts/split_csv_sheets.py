#!/usr/bin/env python3
"""
Split a multi-sheet workbook into one CSV per sheet.

Supports:
  • .xlsx / .xls  — reads every tab directly via openpyxl (recommended)
  • .csv           — falls back to heuristic boundary detection for single-file
                     CSV exports where sheets were concatenated with blank rows

Usage:
    python scripts/split_csv_sheets.py <input.xlsx|input.csv> [output_dir]

If output_dir is omitted, files are written alongside the input file.
Output filenames: <workbook-slug>--<sheet-slug>.csv
"""

import csv
import re
import sys
from pathlib import Path


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def slugify(text: str) -> str:
    """Convert arbitrary text to a safe lowercase filename stem."""
    text = text.lower()
    text = re.sub(r"[^\w\s-]", "", text, flags=re.UNICODE)
    text = re.sub(r"[\s_]+", "-", text.strip())
    text = re.sub(r"-+", "-", text)
    return text.strip("-")[:60]


def write_sheet(rows: list[list], out_path: Path) -> None:
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerows(rows)


# ---------------------------------------------------------------------------
# XLSX path — one tab per file, preserving all sheets
# ---------------------------------------------------------------------------

def split_xlsx(input_path: Path, output_dir: Path) -> list[Path]:
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    output_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(input_path, data_only=True)
    base_stem = slugify(input_path.stem)
    output_files: list[Path] = []

    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([("" if cell is None else str(cell)) for cell in row])

        # Drop trailing fully-blank rows
        while rows and all(c == "" for c in rows[-1]):
            rows.pop()

        slug = slugify(sheet_name) or f"sheet-{idx + 1}"
        out_path = output_dir / f"{base_stem}--{slug}.csv"
        write_sheet(rows, out_path)
        print(f"  [{idx + 1}] {sheet_name!r:30s} → {out_path.name}  ({len(rows)} rows)")
        output_files.append(out_path)

    return output_files


# ---------------------------------------------------------------------------
# CSV fallback — heuristic boundary detection
# ---------------------------------------------------------------------------

def _is_title_row(row: list[str]) -> bool:
    """A row is a sheet title if only the first cell is non-empty."""
    if not row or not row[0].strip():
        return False
    return all(cell.strip() == "" for cell in row[1:])


def _is_blank(row: list[str]) -> bool:
    return all(cell.strip() == "" for cell in row)


def split_csv(input_path: Path, output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)

    with input_path.open(newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.reader(fh))

    boundaries: list[int] = []
    for i, row in enumerate(rows):
        if _is_title_row(row):
            if i == 0 or all(_is_blank(rows[j]) for j in range(max(0, i - 2), i)):
                boundaries.append(i)

    if not boundaries:
        print(f"No sheet boundaries detected in {input_path.name}.")
        print("If this file has multiple tabs, download it as .xlsx instead.")
        return []

    base_stem = slugify(input_path.stem)
    output_files: list[Path] = []

    for idx, start in enumerate(boundaries):
        end = boundaries[idx + 1] if idx + 1 < len(boundaries) else len(rows)
        title = rows[start][0].strip()
        slug = slugify(title) or f"sheet-{idx + 1}"
        out_path = output_dir / f"{base_stem}--{slug}.csv"

        sheet_rows = rows[start:end]
        while sheet_rows and _is_blank(sheet_rows[-1]):
            sheet_rows.pop()

        write_sheet(sheet_rows, out_path)
        print(f"  [{idx + 1}] {title!r:40s} → {out_path.name}  ({len(sheet_rows)} rows)")
        output_files.append(out_path)

    return output_files


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        print(f"Error: {input_path} not found")
        sys.exit(1)

    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else input_path.parent

    print(f"Splitting: {input_path.name}\n")

    suffix = input_path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        files = split_xlsx(input_path, output_dir)
    elif suffix == ".csv":
        print("Note: CSV mode uses heuristic boundary detection.")
        print("For full fidelity, download the source workbook as .xlsx.\n")
        files = split_csv(input_path, output_dir)
    else:
        print(f"Unsupported file type: {suffix}")
        sys.exit(1)

    print(f"\nWrote {len(files)} sheet file(s) to {output_dir}")


if __name__ == "__main__":
    main()
